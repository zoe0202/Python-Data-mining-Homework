import openpyxl
import json


def read(file):
    with open(file, 'r', encoding='utf8') as f:
        data1 = json.load(f)
        data = data1["data"]
    f.close()
    return data


def write(file):
    context = read(file)
    excel = openpyxl.Workbook()
    sheet1 = excel.create_sheet('sheet1', index=0)
    sheet1['A1'] = '类别'
    sheet1['B1'] = '发布时间'
    sheet1['C1'] = '发布者'
    sheet1['D1'] = '弹幕内容'

    for i in range(0, len(context)):
        for j in range(0, 4):
            sheet1.cell(row=i + 2, column=j + 1).value = context[i][j]
            # 因为有表头，所以要加2.

    excel.save(r'C:\Users\Surface\Desktop\result-danmu.xlsx')


if __name__ == '__main__':
    write(r'C:\Users\Surface\Desktop\聆听丶芒果鱼直播间时间切片弹幕.json')
