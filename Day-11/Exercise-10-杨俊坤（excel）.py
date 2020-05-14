''''
将当前文件夹下的“聆听丶芒果鱼直播间时间切片弹幕.json”转存为Excel格式，转存要求如下： 将列表中的每一个元素(一条弹幕信息)转存为Excel中的一行，并在Excel中增加表头一行，各列的名称分别为“类别”、“发布时间”、“发布者”、“弹幕内容”。
'''
import json
from openpyxl import Workbook
from openpyxl import load_workbook

def danmu():
    try:
        list=[]                    #建立一个储存字符串的列表
        with open(r'C:\Users\NealPayne\Desktop\聆听丶芒果鱼直播间时间切片弹幕.json', 'r',encoding='utf-8') as f:         #打开json文件
            file=json.load(f)               #运用load函数将json转为load函数
            data=file['data']             #用【'data'】是提取文件内'data'后的内容
        for i in data:                       #将data的内容循环写入list中
            list.append(i)
        wb = load_workbook(r'C:\Users\NealPayne\Desktop\弹幕.xlsx')            #用load_workbook读取excel表格
        ws = wb.create_sheet("弹幕", 0)                                                                #定义、读取表名
        row_1=['类别','发布时间','发布者','弹幕内容']                                               #用row定义列表应该比写四个函数赋值方便
        ws.append(row_1)
        for x in list:                                                                                                #sheet.append循环写入excel
            ws.append(x)                                                                                         #     ！这里不能写成ws.append(list) 不然会直接报错！
        wb.save(r'C:\Users\NealPayne\Desktop\弹幕.xlsx')                                  #别忘了保存额
    except FileNotFoundError as e:
        print('指定的文件无法打开.')
    except IOError as e:
        print('读写文件时出现错误.')
    print('操作完成！')
if __name__ == '__main__':
    danmu()

#我觉得用代码调整excel的格式不如手动在excel内调整方便 所以就没在这里调整格式了

