

import openpyxl
import json


def readFromJson(file):
    with open("萝莉酱直播间时间切片弹幕.json", 'r', encoding='utf8') as fr:
        jsonData = json.load(fr)
        return jsonData



def writeToExcel(file):
    jn = readFromJson(file)
    excl = openpyxl.Workbook()
    sheet1 = excl.create_sheet('sheet1', index=0)
    sheet1.append(["类别","发布时间","发布者","弹幕内容","弹幕包含颜文字"])
    for c, i in enumerate(jn[0].keys()):
        sheet1.cell(row=1, column=c + 1, value=i)
        sheet1.append(i)
    for r, i in enumerate(jn):
        row = r + 2
        for c, d in enumerate(sheet1):
            sheet1.cell(row=row, column=c + 1, value=i.get(d, ""))

    excl.save("萝莉酱直播间时间切片弹幕 .xlsx")


