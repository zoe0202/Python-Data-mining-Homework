import json
import openpyxl
f =open('D:\\python\\zhouxuruixiaopengyou\\Python-Data-mining-Tutorial\\Week-02\\萝莉酱直播间时间切片弹幕.json',encoding='utf-8')
dt=f.read()
jsonData = json.loads(dt)
print(jsonData['data'])
def writeToExcel():
    excel=openpyxl.Workbook()
    worksheet=excel.create_sheet("mysheet",index=0)
    directionCell = worksheet.cell(row=1, column=1)
    directionCell.value = "类别"
    directionCell = worksheet.cell(row=1, column=2)
    directionCell.value = "发布时间"
    directionCell = worksheet.cell(row=1, column=3)
    directionCell.value = "发布者"
    directionCell = worksheet.cell(row=1, column=4)
    directionCell.value = "弹幕内容"
    c=0
    while c < len(jsonData['data'])-1:
        c=c+1
        directionCell = worksheet.cell(row=c+1, column=1)
        directionCell.value = jsonData['data'][c-1][0]
        directionCell = worksheet.cell(row=c+1, column=2)
        directionCell.value = jsonData['data'][c-1][1]
        directionCell = worksheet.cell(row=c+1, column=3)
        directionCell.value = jsonData['data'][c-1][2]
        directionCell = worksheet.cell(row=c+1, column=4)
        directionCell.value = jsonData['data'][c-1][3]

    excel.save("D:\\test.xlsx")
writeToExcel()
print("操作完成")