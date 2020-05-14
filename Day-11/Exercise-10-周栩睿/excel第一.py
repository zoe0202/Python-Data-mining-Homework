import json
import openpyxl
def read_json(file):
    with open('D:\\python\\zhouxuruixiaopengyou\\Python-Data-mining-Tutorial\\Week-02\\萝莉酱直播间时间切片弹幕.json', "r",encoding='utf-8') as f:
        data = f.read()
        json_data = json.loads(data)
    return json_data
def writeToExcel(file):
    json=read_json(file)
    s='json'
    b=s.split(',')
    excel=openpyxl.Workbook()
    worksheet=excel.create_sheet("mysheet",index=0)
    directionCell = worksheet.cell(row=1, column=1)
    directionCell.value = "类别"
    directionCell = worksheet.cell(row=1, column=2)
    directionCell.value = "发布时间"
    directionCell = worksheet.cell(row=1, column=3)
    directionCell.value = "发布者"
    directionCell = worksheet.cell(row=1, column=4)
    directionCell.value = "弹幕内容"
    for a in range(len(b)):
        for i in range(4):
            worksheet.cell(row=a+1, column=i).value=json[i]
    excel.save("D:\\python\\test.xlsx")
writeToExcel("D:\\python\\test.xlsx")
print("操作完成")