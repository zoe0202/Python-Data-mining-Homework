import json
import openpyxl


def read_json(file):
    with open('D:\\python\\zhouxuruixiaopengyou\\Python-Data-mining-Tutorial\\Week-02\\萝莉酱直播间时间切片弹幕.json', "r",
              encoding='utf-8') as f:
        data = f.read()
        json_data = json.loads(data)
    return json_data


def writeToExcel(file):
    json = read_json(file)
    excel = openpyxl.Workbook()
    sheet = excel.create_sheet("mysheet", index=0)
    directionCell = sheet.cell(row=1, column=1)
    directionCell.value = "类别"
    directionCell = sheet.cell(row=1, column=2)
    directionCell.value = "发布时间"
    directionCell = sheet.cell(row=1, column=3)
    directionCell.value = "发布者"
    directionCell = sheet.cell(row=1, column=4)
    directionCell.value = "弹幕内容"
    excel.save('D:\\python\\test.xlsx')


writeToExcel('D:\\python\\test.xlsx')
