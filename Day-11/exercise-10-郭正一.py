# """
# 目的：将弹幕json转化为excel
# 作者：徐昭
# """
# import openpyxl
# import json
#
# # 新建表格
# wb = openpyxl.Workbook()
# sheet = wb.create_sheet()
# # 美化表格
# sheet.column_dimensions['A'].width = 10
# sheet.column_dimensions['B'].width = 20
# sheet.column_dimensions['C'].width = 20
# sheet.column_dimensions['D'].width = 50
# # 表头赋值
# sheet['A1'] = '类别'
# sheet['B1'] = '发布时间'
# sheet['C1'] = '发布者'
# sheet['D1'] = '弹幕内容'
# json_path = 'E:\\Python\\danmu.json'
# # 读取JSON文件
# with open(json_path, 'r', encoding='utf8') as f:
#     danmu = json.load(f)
#     data = danmu["data"]
#     # 循环填充表头下方的表格
#     for j in range(0, len(data)):
#         for id in range(0, 4):
#             sheet[chr(ord('A') + id) + str(j + 2)] = data[j][id]
# # 保存文件
# wb.save('E:\\Python\\danmu.xlsx')

import openpyxl
import json
#前期准备

from openpyxl import Workbook
wb = Workbook()
ws1 = wb.create_sheet()
#第一步，建立工作簿
cell1 = ws1["A1"]
cell2 = ws1['B1']
cell3 = ws1['C1']
cell4 = ws1['D1']
cell1.value = '类别'
cell2.value = '发布时间'
cell3.value = '发布者'
cell4.value = '弹幕内容'
#第二步，设定表头
json_path = 'E:\\json\\mangguo.json'
with open(json_path, 'r', encoding='utf-8') as f:
    Danmu = json.load(f)
    data = Danmu['data']
#第三步，读取json文件
for x in range(0,len(data)):
    for y in range(0,4):
        ws1[chr(ord('A') + y) + str(x + 2)] = data[x][y]
#第四步，写入
wb.save('E:\\json\\mangguo.xlsx')
#第五步，保存



